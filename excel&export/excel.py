# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
 @Time : 2019/12/3 14:10
 @Auth : 明明
 @IDE  : PyCharm
 """
import platform
from urllib import parse

import openpyxl
import pymongo
from openpyxl import Workbook


def combine_sheet():
    """
    获取多个sheet内容
    :return:
    """
    wb = openpyxl.load_workbook("20191104_tm_kf.xlsx")

    # 获取workbook中所有的表格
    sheets = wb.get_sheet_names()

    print(sheets)

    list1 = []
    # 循环遍历所有sheet
    for t in range(len(sheets)):
        sheet = wb.get_sheet_by_name(sheets[t])

        print('\n\n第' + str(t + 1) + '个sheet: ' + sheet.title + '->>>')

        # len_row代表表中有多少行数据，len_column代表excel表中有多少列数据
        len_row = sheet.max_row
        len_column = sheet.max_column
        # 合并的时候只保留第一张表的表头部分
        if t == 0:
            for i in range(1, len_row + 1):
                list2 = []
                for j in range(1, len_column + 1):
                    list2.append(sheet.cell(row=i, column=j).value)
                list1.append(list2)
        else:
            for i in range(2, len_row + 1):
                list2 = []
                for j in range(1, len_column + 1):
                    list2.append(sheet.cell(row=i, column=j).value)
                list1.append(list2)


def generate_sheet():
    from openpyxl import Workbook

    wb = Workbook()

    ws1 = wb.create_sheet('sheet1', 0)
    ws2 = wb.create_sheet('sheet2', 1)
    ws3 = wb.create_sheet('sheet3', 2)

    ws1['A1'] = ws1.title
    ws2['A1'] = ws2.title
    ws3['A1'] = ws3.title

    ws1.sheet_properties.tabColor = '1072BA'
    ws2.sheet_properties.tabColor = '1072BA'
    ws3.sheet_properties.tabColor = '1072BA'

    wb.save('d:\\sample.xlsx')


def export_jd():
    if 'Win' in platform.system():
        client = pymongo.MongoClient('127.0.0.1', 27017)
    else:
        client = pymongo.MongoClient('10.1.6.173', 28018)

    db = client["voc"]
    collection = db["final_info"]

    wb = Workbook()
    ws1 = wb.create_sheet('商城信息', 0)
    ws2 = wb.create_sheet('官网配件信息', 1)
    for info in collection.find({"plat": {"$ne": "tm"}}).sort("waiter"):
        info.pop("_id")
        cols_val = []
        for k in info:
            cols_val.append(repr(info[k]))
        ws1.append(cols_val)  # 写值

    for info in collection.find({"plat": {"$ne": "tm"}}).sort("waiter"):
        info.pop("_id")
        cols_val = []
        for k in info:
            cols_val.append(repr(info[k]))
        ws2.append(cols_val)  # 写值

    wb.save("20191104_tm_kf.xlsx")
    print("存储完毕")


def parse_dt(li):
    li_new = []
    for info in li:
        dt = {}
        for key in info.keys():
            value = info[key]
            if isinstance(value, dict):
                for sub_key in value.keys():
                    if sub_key == "url":
                        dt[f"{key}.{sub_key}"] = parse.urljoin("https://", value[sub_key])
                    else:
                        dt[f"{key}.{sub_key}"] = value[sub_key]
            else:
                dt[key] = value
        li_new.append(dt)
    return li_new


def parse_dict(info):
    if isinstance(info, dict):
        keys = [key for key in info.keys()]
        for i in keys:
            if isinstance(info[i], dict):
                deal_dt = info.pop(i)
                deal_new = {f"{i}-{_}": deal_dt[_] for _ in deal_dt.keys()}
                deal_new = parse_dict(deal_new)
                info.update(deal_new)
            elif isinstance(info[i], list):
                deal_li = info.pop(i)
                deal_new = {f"{i}-{_[0]}": _[1] for _ in enumerate(deal_li)}
                deal_new = parse_dict(deal_new)
                info.update(deal_new)
            else:
                continue
        return info
    elif isinstance(info, list):
        deal_dt = {str(_[0]): _[1] for _ in enumerate(info)}
        deal_dt = parse_dict(deal_dt)
        return deal_dt
    else:
        return {}


if __name__ == '__main__':
    # export()
    # generate_sheet()
    dt = {"one": {"two": {"three": "hong"}}}
    res = parse_dict(dt)
    print(res)
