# !/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
 @Time : 2019/12/3 14:10
 @Auth : 明明
 @IDE  : PyCharm
 """

import pymongo
from openpyxl import Workbook


def test():
    print(1)


def export_tm():
    client = pymongo.MongoClient('127.0.0.1', 27017)

    db = client["voc"]
    collection = db["final_info"]

    wb = Workbook()
    ws1 = wb.create_sheet('商城信息', 0)
    ws2 = wb.create_sheet('官网配件信息', 1)

    i = 1
    for info in collection.find({"plat": {"$eq": "tm"}}).sort("sku", 1).sort("keyword", 1).sort("brand", 1):
        info.pop("_id")
        info.pop("parts")
        # base = info.pop("base")
        # info['shop_name'] = base.get("shop_name")
        # price = info.pop("price")
        # info['price'] = price.get("price")

        if i == 1:
            cols_val = []
            for k in info:
                cols_val.append(repr(k))
            ws1.append(cols_val)  # 写值

        cols_val = []
        for k in info:
            cols_val.append(repr(info[k]))
        ws1.append(cols_val)  # 写值

        i += 1

    j = 1
    for info in collection.find({"plat": {"$eq": "tm"}}).sort("sku", 1).sort("keyword", 1).sort("brand", 1):
        info.pop("_id")
        item = {
            "brand": info['brand'],
            "sku": info['sku'],
            "parts": info['parts']
        }
        if j == 1:
            cols_val = []
            for k in item:
                cols_val.append(repr(k))
            ws2.append(cols_val)  # 写值

        cols_val = []
        for k in item:
            cols_val.append(repr(item[k]))
        ws2.append(cols_val)  # 写值

        j += 1

    wb.save("20191104_tm_kf.xlsx")
    print("存储完毕")


if __name__ == '__main__':
    # test()
    export_tm()
